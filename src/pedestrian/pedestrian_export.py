"""Excel export for pedestrian counter results."""

import logging

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.common.survey_widget import SurveyInfo

logger = logging.getLogger(__name__)


def export_pedestrian_results(
    results: dict, survey: SurveyInfo, filepath: str
) -> None:
    """Export pedestrian counter results to a formatted Excel file."""
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        _write_summary_sheet(writer, results)
        _write_per_line_sheet(writer, results)
        _write_interval_sheet(writer, results)

    _format_workbook(filepath, survey, results)
    logger.info(
        "Pedestrian results exported to %s (%d pedestrians)",
        filepath,
        results.get("grand_total", 0),
    )


def _write_summary_sheet(writer, results: dict):
    """Grand totals aggregated across all lines."""
    per_line = results.get("per_line", {})
    line_labels = results.get("line_labels", list(per_line.keys()))

    rows = []
    total_in = 0
    total_out = 0
    for label in line_labels:
        line_data = per_line.get(label, {})
        count_in = line_data.get("count_in", 0)
        count_out = line_data.get("count_out", 0)
        total_in += count_in
        total_out += count_out
        rows.append({
            "Count Line": label,
            "Pedestrians (In)": count_in,
            "Pedestrians (Out)": count_out,
            "Total": count_in + count_out,
        })

    rows.append({
        "Count Line": "TOTAL",
        "Pedestrians (In)": total_in,
        "Pedestrians (Out)": total_out,
        "Total": total_in + total_out,
    })

    pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="Summary")


def _write_per_line_sheet(writer, results: dict):
    """Breakdown by count line and direction."""
    per_line = results.get("per_line", {})
    line_labels = results.get("line_labels", list(per_line.keys()))

    rows = []
    for label in line_labels:
        line_data = per_line.get(label, {})
        for direction in ["IN", "OUT"]:
            key = f"count_{direction.lower()}"
            count = line_data.get(key, 0)
            if count > 0:
                rows.append({
                    "Count Line": label,
                    "Direction": direction,
                    "Pedestrian Count": count,
                })

        line_total = line_data.get("total", 0)
        if line_total > 0:
            rows.append({
                "Count Line": label,
                "Direction": "SUBTOTAL",
                "Pedestrian Count": line_total,
            })

    if not rows:
        rows = [{"Count Line": "", "Direction": "", "Pedestrian Count": 0}]

    pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="Per Line")


def _write_interval_sheet(writer, results: dict):
    """15-minute interval binning per line."""
    per_line = results.get("per_line", {})
    line_labels = results.get("line_labels", list(per_line.keys()))

    all_intervals: set[str] = set()
    for line_data in per_line.values():
        intervals = line_data.get("intervals", {})
        all_intervals.update(intervals.keys())

    all_intervals_sorted = sorted(all_intervals)

    rows = []
    for interval_key in all_intervals_sorted:
        for label in line_labels:
            line_data = per_line.get(label, {})
            intervals = line_data.get("intervals", {})
            interval_data = intervals.get(interval_key, {})

            for direction in ["in", "out"]:
                count = interval_data.get(direction, 0)
                rows.append({
                    "Time Interval": interval_key,
                    "Count Line": label,
                    "Direction": direction.upper(),
                    "Pedestrian Count": count,
                })

    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(
            columns=["Time Interval", "Count Line", "Direction", "Pedestrian Count"]
        )

    df.to_excel(writer, index=False, sheet_name="Interval Data")


def _format_workbook(filepath: str, survey: SurveyInfo, results: dict):
    """Apply formatting and survey header to the workbook."""
    wb = load_workbook(filepath)

    header_fill = PatternFill(
        start_color="0f3460", end_color="0f3460", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin", color="2a2a4a"),
        right=Side(style="thin", color="2a2a4a"),
        top=Side(style="thin", color="2a2a4a"),
        bottom=Side(style="thin", color="2a2a4a"),
    )

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, max_col=ws.max_column
        ):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Add survey info to summary sheet
    ws = wb["Summary"]
    header_rows = [
        (
            "Matrix Pedestrian Counter Report",
            Font(bold=True, size=14, color="e94560"),
        ),
    ]
    if survey.job_number or survey.job_name:
        header_rows.append((
            f"Job: {survey.job_number}  —  {survey.job_name}",
            Font(size=11),
        ))
    if survey.site_number or survey.site_name:
        header_rows.append((
            f"Site: {survey.site_number}  —  {survey.site_name}",
            Font(size=11),
        ))
    if survey.camera_number:
        header_rows.append((f"Camera: {survey.camera_number}", Font(size=11)))

    duration = results.get("duration_sec", 0)
    mins = int(duration // 60)
    secs = int(duration % 60)
    header_rows.append((f"Video Duration: {mins}m {secs}s", Font(size=11)))
    header_rows.append((
        f"Total Pedestrians: {results.get('grand_total', 0)}", Font(size=11),
    ))

    line_labels = results.get("line_labels", [])
    header_rows.append((f"Count Lines: {', '.join(line_labels)}", Font(size=11)))

    video_count = results.get("video_count", 1)
    if video_count > 1:
        header_rows.append((f"Videos Processed: {video_count}", Font(size=11)))

    ws.insert_rows(1, len(header_rows) + 1)
    for i, (text, font) in enumerate(header_rows):
        ws[f"A{i + 1}"] = text
        ws[f"A{i + 1}"].font = font

    wb.save(filepath)
