"""Excel export for midblock vehicle counter results with per-line data and speed."""

import logging

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.counter.vehicle_classifier import AUSTROADS_CLASSES
from src.common.survey_widget import SurveyInfo

logger = logging.getLogger(__name__)


def export_counter_results(
    results: dict, survey: SurveyInfo, filepath: str
) -> None:
    """Export vehicle counter results to a formatted Excel file.

    Args:
        results: Dict with keys: per_line, grand_total, fps, duration_sec, etc.
        survey: Survey details for the report header.
        filepath: Output .xlsx file path.
    """
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        _write_summary_sheet(writer, results)
        _write_per_line_sheet(writer, results)
        _write_interval_sheet(writer, results)
        if results.get("speed_enabled"):
            _write_speed_sheet(writer, results)

    _format_workbook(filepath, survey, results)
    logger.info(
        "Counter results exported to %s (%d vehicles)",
        filepath,
        results.get("grand_total", 0),
    )


def _write_summary_sheet(writer, results: dict):
    """Grand totals aggregated across all lines."""
    per_line = results.get("per_line", {})

    all_in: dict[str, int] = {}
    all_out: dict[str, int] = {}
    for line_data in per_line.values():
        for cls, count in line_data.get("counts_in", {}).items():
            all_in[cls] = all_in.get(cls, 0) + count
        for cls, count in line_data.get("counts_out", {}).items():
            all_out[cls] = all_out.get(cls, 0) + count

    all_classes = sorted(set(all_in.keys()) | set(all_out.keys()))
    if not all_classes:
        all_classes = ["1"]

    rows = []
    for cls_key in all_classes:
        info = AUSTROADS_CLASSES.get(cls_key, {"name": f"Class {cls_key}", "code": cls_key})
        count_in = all_in.get(cls_key, 0)
        count_out = all_out.get(cls_key, 0)
        rows.append({
            "Class": cls_key,
            "Code": info["code"],
            "Description": info["name"],
            "Count (In)": count_in,
            "Count (Out)": count_out,
            "Total": count_in + count_out,
        })

    total_in = sum(all_in.values())
    total_out = sum(all_out.values())
    rows.append({
        "Class": "",
        "Code": "",
        "Description": "TOTAL",
        "Count (In)": total_in,
        "Count (Out)": total_out,
        "Total": total_in + total_out,
    })

    pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="Summary")


def _write_per_line_sheet(writer, results: dict):
    """Breakdown by count line, class, and direction."""
    per_line = results.get("per_line", {})
    line_labels = results.get("line_labels", list(per_line.keys()))

    rows = []
    for label in line_labels:
        line_data = per_line.get(label, {})
        counts_in = line_data.get("counts_in", {})
        counts_out = line_data.get("counts_out", {})
        all_classes = sorted(set(counts_in.keys()) | set(counts_out.keys()))

        for cls_key in all_classes:
            info = AUSTROADS_CLASSES.get(cls_key, {"name": f"Class {cls_key}", "code": cls_key})
            for direction, counts in [("IN", counts_in), ("OUT", counts_out)]:
                count = counts.get(cls_key, 0)
                if count > 0:
                    rows.append({
                        "Count Line": label,
                        "Class": cls_key,
                        "Code": info["code"],
                        "Description": info["name"],
                        "Direction": direction,
                        "Count": count,
                    })

        line_total = line_data.get("total", 0)
        if line_total > 0:
            rows.append({
                "Count Line": label,
                "Class": "",
                "Code": "",
                "Description": f"SUBTOTAL ({label})",
                "Direction": "",
                "Count": line_total,
            })

    if not rows:
        rows = [{"Count Line": "", "Class": "", "Code": "", "Description": "No data", "Direction": "", "Count": 0}]

    pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="Per Line")


def _write_interval_sheet(writer, results: dict):
    """15-minute interval binning per line."""
    per_line = results.get("per_line", {})
    line_labels = results.get("line_labels", list(per_line.keys()))

    all_intervals: set[str] = set()
    all_classes: set[str] = set()
    for line_data in per_line.values():
        intervals = line_data.get("intervals", {})
        all_intervals.update(intervals.keys())
        for interval_data in intervals.values():
            all_classes.update(interval_data.keys())

    all_intervals_sorted = sorted(all_intervals)
    all_classes_sorted = sorted(all_classes) or ["1"]

    rows = []
    for interval_key in all_intervals_sorted:
        for label in line_labels:
            line_data = per_line.get(label, {})
            intervals = line_data.get("intervals", {})
            interval_data = intervals.get(interval_key, {})

            for direction in ["in", "out"]:
                row: dict = {
                    "Time Interval": interval_key,
                    "Count Line": label,
                    "Direction": direction.upper(),
                }
                interval_total = 0
                for cls_key in all_classes_sorted:
                    cls_info = AUSTROADS_CLASSES.get(cls_key, {"code": cls_key})
                    col_name = f"Class {cls_key} ({cls_info['code']})"
                    count = interval_data.get(cls_key, {}).get(direction, 0)
                    row[col_name] = count
                    interval_total += count
                row["Total"] = interval_total
                rows.append(row)

    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=["Time Interval", "Count Line", "Direction", "Total"])

    df.to_excel(writer, index=False, sheet_name="Interval Data")


def _write_speed_sheet(writer, results: dict):
    """Speed data aggregated by interval and line."""
    per_line = results.get("per_line", {})
    line_labels = results.get("line_labels", list(per_line.keys()))

    rows = []
    for label in line_labels:
        line_data = per_line.get(label, {})
        speeds = line_data.get("speeds", [])

        if not speeds:
            continue

        # Group speeds by interval
        by_interval: dict[str, list] = {}
        for s in speeds:
            key = s.get("interval", "unknown")
            by_interval.setdefault(key, []).append(s)

        for interval_key in sorted(by_interval.keys()):
            interval_speeds = by_interval[interval_key]
            speed_vals = [s["speed_kmh"] for s in interval_speeds]

            if not speed_vals:
                continue

            rows.append({
                "Time Interval": interval_key,
                "Count Line": label,
                "Vehicle Count": len(speed_vals),
                "Avg Speed (km/h)": round(sum(speed_vals) / len(speed_vals), 1),
                "85th %ile Speed (km/h)": round(
                    float(np.percentile(speed_vals, 85)), 1
                ),
                "Max Speed (km/h)": round(max(speed_vals), 1),
                "Min Speed (km/h)": round(min(speed_vals), 1),
            })

    if not rows:
        rows = [{"Time Interval": "", "Count Line": "", "Vehicle Count": 0,
                 "Avg Speed (km/h)": 0, "85th %ile Speed (km/h)": 0,
                 "Max Speed (km/h)": 0, "Min Speed (km/h)": 0}]

    pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="Speed Data")


def _format_workbook(filepath: str, survey: SurveyInfo, results: dict):
    """Apply formatting and survey header to the workbook."""
    wb = load_workbook(filepath)

    header_fill = PatternFill(start_color="0f3460", end_color="0f3460", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin", color="2a2a4a"),
        right=Side(style="thin", color="2a2a4a"),
        top=Side(style="thin", color="2a2a4a"),
        bottom=Side(style="thin", color="2a2a4a"),
    )

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Add survey info to summary sheet
    ws = wb["Summary"]
    header_rows = [
        ("Matrix Midblock Vehicle Counter Report", Font(bold=True, size=14, color="e94560")),
    ]
    if survey.job_number or survey.job_name:
        header_rows.append((
            f"Job: {survey.job_number}  —  {survey.job_name}",
            Font(size=11),
        ))
    if survey.site_number or survey.site_name:
        header_rows.append((
            f"Site: {survey.site_number}  —  {survey.site_name}",
            Font(size=11),
        ))
    if survey.camera_number:
        header_rows.append((f"Camera: {survey.camera_number}", Font(size=11)))

    duration = results.get("duration_sec", 0)
    mins = int(duration // 60)
    secs = int(duration % 60)
    header_rows.append((f"Video Duration: {mins}m {secs}s", Font(size=11)))
    header_rows.append((f"Total Vehicles: {results.get('grand_total', 0)}", Font(size=11)))

    line_labels = results.get("line_labels", [])
    header_rows.append((f"Count Lines: {', '.join(line_labels)}", Font(size=11)))

    video_count = results.get("video_count", 1)
    if video_count > 1:
        header_rows.append((f"Videos Processed: {video_count}", Font(size=11)))

    ws.insert_rows(1, len(header_rows) + 1)
    for i, (text, font) in enumerate(header_rows):
        ws[f"A{i + 1}"] = text
        ws[f"A{i + 1}"].font = font

    wb.save(filepath)
