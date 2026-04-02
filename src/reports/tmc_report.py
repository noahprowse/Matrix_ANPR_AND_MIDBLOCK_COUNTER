"""Matrix Traffic Data — Intersection Turning Movement Count (TMC) Report.

Generates a branded Excel workbook with TMC Summary, TMC 15-Min, TMC Hourly,
and O-D Matrix sheets for intersection counting jobs.
"""

from __future__ import annotations

from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.common.data_models import AUSTROADS_CLASSES, JobConfig, SiteConfig
from src.reports.matrix_branding import (
    ACCENT_BLUE,
    HEADER_BG,
    apply_data_borders,
    apply_header_style,
    auto_size_columns,
    get_border,
    get_header_fill,
    get_header_font,
    get_subheader_fill,
    get_total_fill,
)
from src.reports.report_engine import BaseReport

# Standard turning movements
MOVEMENTS: list[str] = ["Left", "Through", "Right", "U-Turn"]


class TMCReport(BaseReport):
    """Intersection Turning Movement Count report.

    Expected ``tmc_data`` structure passed to ``generate()``::

        {
            "approaches": ["North", "South", "East", "West"],
            "class_codes": ["1", "1M", "3", ...],
            "movements": ["Left", "Through", "Right", "U-Turn"],
            "summary": {
                "North": {
                    "Left":    {"1": 50, "1M": 2, ...},
                    "Through": {"1": 200, "1M": 5, ...},
                    "Right":   {"1": 80, "1M": 1, ...},
                    "U-Turn":  {"1": 5, "1M": 0, ...},
                },
                ...
            },
            "interval_data": [
                {
                    "interval": "07:00",
                    "counts": {
                        "North": {
                            "Left":    {"1": 5, "1M": 0, ...},
                            "Through": {"1": 20, "1M": 1, ...},
                            ...
                        },
                        ...
                    },
                },
                ...
            ],
            "hourly_data": [  ...same structure as interval_data...  ],
        }

    Expected ``od_matrix`` structure::

        {
            "origins": ["North", "South", "East", "West"],
            "destinations": ["North", "South", "East", "West"],
            "matrix": {
                "North": {"North": 0, "South": 120, "East": 80, "West": 50},
                ...
            },
        }
    """

    def __init__(self, job_config: JobConfig) -> None:
        super().__init__(job_config)

    def generate(
        self,
        tmc_data: dict[str, Any],
        od_matrix: dict[str, Any],
        site_config: SiteConfig,
        approach_config: Optional[dict[str, Any]] = None,
    ) -> None:
        """Build the full TMC report workbook.

        Parameters
        ----------
        tmc_data : dict
            Turning movement count data (see class docstring).
        od_matrix : dict
            Origin-destination matrix data.
        site_config : SiteConfig
            Site-level configuration.
        approach_config : dict | None
            Optional approach metadata (compass directions, lane info, etc.).
        """
        wb = self._create_workbook()

        approaches: list[str] = tmc_data.get("approaches", [])
        class_codes: list[str] = tmc_data.get("class_codes", [])
        movements: list[str] = tmc_data.get("movements", MOVEMENTS)
        summary: dict[str, dict[str, dict[str, int]]] = tmc_data.get("summary", {})
        interval_data: list[dict[str, Any]] = tmc_data.get("interval_data", [])
        hourly_data: list[dict[str, Any]] = tmc_data.get("hourly_data", [])

        self._build_summary_sheet(wb, approaches, movements, class_codes, summary)
        self._build_15min_sheet(wb, interval_data, approaches, movements, class_codes)
        self._build_hourly_sheet(wb, hourly_data, approaches, movements, class_codes)
        self._build_od_sheet(wb, od_matrix)

    # ── Sheet builders ──

    def _build_summary_sheet(
        self,
        wb: Workbook,
        approaches: list[str],
        movements: list[str],
        class_codes: list[str],
        summary: dict[str, dict[str, dict[str, int]]],
    ) -> None:
        """Sheet 'TMC Summary': matrix format — rows=approach, cols=movement, split by class."""
        ws = self._add_sheet(wb, "TMC Summary")
        row = self._write_header(ws, "Turning Movement Count Summary")

        # Build headers: Approach | Movement | Class1 | Class2 | ... | Total
        headers = ["Approach", "Movement"]
        for cc in class_codes:
            class_info = AUSTROADS_CLASSES.get(cc, {})
            headers.append(class_info.get("name", cc))
        headers.append("Total")

        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=row, column=col_idx, value=h)
        apply_header_style(ws, row, 1, len(headers))

        data_start = row + 1
        current_row = data_start
        total_fill = get_total_fill()
        subheader_fill = get_subheader_fill()
        border = get_border()

        for approach in approaches:
            approach_start_row = current_row
            approach_movements = summary.get(approach, {})

            for mvt in movements:
                ws.cell(row=current_row, column=1, value=approach)
                ws.cell(row=current_row, column=2, value=mvt)

                mvt_counts = approach_movements.get(mvt, {})
                mvt_total = 0
                for col_idx, cc in enumerate(class_codes, start=3):
                    val = mvt_counts.get(cc, 0)
                    ws.cell(row=current_row, column=col_idx, value=val)
                    mvt_total += val
                ws.cell(row=current_row, column=len(headers), value=mvt_total)
                current_row += 1

            # Approach subtotal row
            ws.cell(row=current_row, column=1, value=approach)
            ws.cell(row=current_row, column=2, value="Subtotal")
            for col_idx, cc in enumerate(class_codes, start=3):
                approach_class_total = sum(
                    approach_movements.get(m, {}).get(cc, 0) for m in movements
                )
                ws.cell(row=current_row, column=col_idx, value=approach_class_total)
            approach_grand = sum(
                sum(approach_movements.get(m, {}).get(cc, 0) for cc in class_codes)
                for m in movements
            )
            ws.cell(row=current_row, column=len(headers), value=approach_grand)

            for c in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=c).fill = subheader_fill
                ws.cell(row=current_row, column=c).border = border
                ws.cell(row=current_row, column=c).font = Font(name="Calibri", bold=True, size=11)
                ws.cell(row=current_row, column=c).alignment = Alignment(horizontal="center")
            current_row += 1

        data_end = current_row - 1
        apply_data_borders(ws, data_start, data_end, 1, len(headers))

        # Grand total row
        ws.cell(row=current_row, column=1, value="GRAND TOTAL")
        ws.cell(row=current_row, column=2, value="")
        overall_grand = 0
        for col_idx, cc in enumerate(class_codes, start=3):
            class_total = sum(
                sum(summary.get(a, {}).get(m, {}).get(cc, 0) for m in movements)
                for a in approaches
            )
            ws.cell(row=current_row, column=col_idx, value=class_total)
            overall_grand += class_total
        ws.cell(row=current_row, column=len(headers), value=overall_grand)

        for c in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=c).fill = total_fill
            ws.cell(row=current_row, column=c).border = border
            ws.cell(row=current_row, column=c).font = Font(name="Calibri", bold=True, size=11)
            ws.cell(row=current_row, column=c).alignment = Alignment(horizontal="center")

        auto_size_columns(ws)

    def _build_15min_sheet(
        self,
        wb: Workbook,
        interval_data: list[dict[str, Any]],
        approaches: list[str],
        movements: list[str],
        class_codes: list[str],
    ) -> None:
        """Sheet 'TMC 15-Min': per-interval TMC with approach and movement columns."""
        ws = self._add_sheet(wb, "TMC 15-Min")
        row = self._write_header(ws, "Turning Movement Count — 15-Minute Intervals")

        # Headers: Time | Approach1-Left | Approach1-Through | ... | Total Approach1 | ...  | Grand Total
        headers: list[str] = ["Time"]
        for approach in approaches:
            for mvt in movements:
                headers.append(f"{approach} {mvt}")
            headers.append(f"Total {approach}")
        headers.append("Grand Total")

        col_end = len(headers)
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=row, column=col_idx, value=h)
        apply_header_style(ws, row, 1, col_end)

        data_start = row + 1
        current_row = data_start

        for record in interval_data:
            col = 1
            ws.cell(row=current_row, column=col, value=record.get("interval", ""))
            col += 1
            grand_total = 0

            for approach in approaches:
                approach_counts = record.get("counts", {}).get(approach, {})
                approach_total = 0
                for mvt in movements:
                    mvt_counts = approach_counts.get(mvt, {})
                    mvt_total = sum(mvt_counts.get(cc, 0) for cc in class_codes)
                    ws.cell(row=current_row, column=col, value=mvt_total)
                    approach_total += mvt_total
                    col += 1
                ws.cell(row=current_row, column=col, value=approach_total)
                grand_total += approach_total
                col += 1

            ws.cell(row=current_row, column=col, value=grand_total)
            current_row += 1

        data_end = current_row - 1
        if interval_data:
            apply_data_borders(ws, data_start, data_end, 1, col_end)

        # Peak identification
        self._annotate_tmc_peaks(ws, interval_data, approaches, movements, class_codes, current_row, col_end, data_start)

        auto_size_columns(ws)

    def _build_hourly_sheet(
        self,
        wb: Workbook,
        hourly_data: list[dict[str, Any]],
        approaches: list[str],
        movements: list[str],
        class_codes: list[str],
    ) -> None:
        """Sheet 'TMC Hourly': hourly aggregation."""
        ws = self._add_sheet(wb, "TMC Hourly")
        row = self._write_header(ws, "Turning Movement Count — Hourly Summary")

        # Same column structure as 15-min
        headers: list[str] = ["Time"]
        for approach in approaches:
            for mvt in movements:
                headers.append(f"{approach} {mvt}")
            headers.append(f"Total {approach}")
        headers.append("Grand Total")

        col_end = len(headers)
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=row, column=col_idx, value=h)
        apply_header_style(ws, row, 1, col_end)

        data_start = row + 1
        current_row = data_start

        for record in hourly_data:
            col = 1
            ws.cell(row=current_row, column=col, value=record.get("interval", ""))
            col += 1
            grand_total = 0

            for approach in approaches:
                approach_counts = record.get("counts", {}).get(approach, {})
                approach_total = 0
                for mvt in movements:
                    mvt_counts = approach_counts.get(mvt, {})
                    mvt_total = sum(mvt_counts.get(cc, 0) for cc in class_codes)
                    ws.cell(row=current_row, column=col, value=mvt_total)
                    approach_total += mvt_total
                    col += 1
                ws.cell(row=current_row, column=col, value=approach_total)
                grand_total += approach_total
                col += 1

            ws.cell(row=current_row, column=col, value=grand_total)
            current_row += 1

        data_end = current_row - 1
        if hourly_data:
            apply_data_borders(ws, data_start, data_end, 1, col_end)

        # Peak hour highlight
        self._annotate_tmc_peaks(ws, hourly_data, approaches, movements, class_codes, current_row, col_end, data_start)

        auto_size_columns(ws)

    def _build_od_sheet(
        self,
        wb: Workbook,
        od_matrix: dict[str, Any],
    ) -> None:
        """Sheet 'O-D Matrix': full origin-destination matrix."""
        ws = self._add_sheet(wb, "O-D Matrix")
        row = self._write_header(ws, "Origin-Destination Matrix")

        origins: list[str] = od_matrix.get("origins", [])
        destinations: list[str] = od_matrix.get("destinations", [])
        matrix: dict[str, dict[str, int]] = od_matrix.get("matrix", {})

        if not origins or not destinations:
            ws.cell(row=row, column=1, value="No O-D data available.")
            return

        # Headers: Origin \ Destination | Dest1 | Dest2 | ... | Total
        headers = ["Origin \\ Dest"]
        headers.extend(destinations)
        headers.append("Total")

        col_end = len(headers)
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=row, column=col_idx, value=h)
        apply_header_style(ws, row, 1, col_end)

        data_start = row + 1
        current_row = data_start
        total_fill = get_total_fill()
        border = get_border()

        col_totals: dict[str, int] = {d: 0 for d in destinations}
        grand_total = 0

        for origin in origins:
            ws.cell(row=current_row, column=1, value=origin)
            row_total = 0
            for col_idx, dest in enumerate(destinations, start=2):
                val = matrix.get(origin, {}).get(dest, 0)
                ws.cell(row=current_row, column=col_idx, value=val)
                row_total += val
                col_totals[dest] += val
            ws.cell(row=current_row, column=col_end, value=row_total)
            grand_total += row_total
            current_row += 1

        data_end = current_row - 1
        apply_data_borders(ws, data_start, data_end, 1, col_end)

        # Total row
        ws.cell(row=current_row, column=1, value="Total")
        for col_idx, dest in enumerate(destinations, start=2):
            ws.cell(row=current_row, column=col_idx, value=col_totals[dest])
        ws.cell(row=current_row, column=col_end, value=grand_total)

        for c in range(1, col_end + 1):
            ws.cell(row=current_row, column=c).fill = total_fill
            ws.cell(row=current_row, column=c).border = border
            ws.cell(row=current_row, column=c).font = Font(name="Calibri", bold=True, size=11)
            ws.cell(row=current_row, column=c).alignment = Alignment(horizontal="center")

        auto_size_columns(ws)

    # ── Peak annotation helper ──

    def _annotate_tmc_peaks(
        self,
        ws: Worksheet,
        data: list[dict[str, Any]],
        approaches: list[str],
        movements: list[str],
        class_codes: list[str],
        next_row: int,
        col_end: int,
        data_start: int,
    ) -> None:
        """Find and highlight peak intervals and annotate the sheet."""
        if not data:
            return

        def _record_total(record: dict[str, Any]) -> int:
            total = 0
            for approach_counts in record.get("counts", {}).values():
                for mvt_counts in approach_counts.values():
                    if isinstance(mvt_counts, dict):
                        total += sum(mvt_counts.values())
                    elif isinstance(mvt_counts, (int, float)):
                        total += mvt_counts
            return total

        # Find single highest interval
        best_idx = 0
        best_vol = 0
        for i, rec in enumerate(data):
            vol = _record_total(rec)
            if vol > best_vol:
                best_vol = vol
                best_idx = i

        self._highlight_peak_row(ws, data_start + best_idx, 1, col_end)

        # Find peak hour (4 consecutive)
        if len(data) >= 4:
            best_hour_start = 0
            best_hour_vol = 0
            for i in range(len(data) - 3):
                window = sum(_record_total(data[j]) for j in range(i, i + 4))
                if window > best_hour_vol:
                    best_hour_vol = window
                    best_hour_start = i

            peak_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            peak_font = Font(name="Calibri", bold=True, size=11, color=ACCENT_BLUE)
            for i in range(best_hour_start, best_hour_start + 4):
                if i != best_idx:
                    for c in range(1, col_end + 1):
                        ws.cell(row=data_start + i, column=c).fill = peak_fill
                        ws.cell(row=data_start + i, column=c).font = peak_font

            # Annotations
            annotation_row = next_row + 1
            ws.cell(row=annotation_row, column=1, value="Peak Interval:").font = Font(
                name="Calibri", bold=True, size=10, color=HEADER_BG
            )
            ws.cell(
                row=annotation_row, column=2,
                value=f"{data[best_idx].get('interval', '')} — {best_vol} vehicles",
            )
            annotation_row += 1
            ws.cell(row=annotation_row, column=1, value="Peak Hour:").font = Font(
                name="Calibri", bold=True, size=10, color=HEADER_BG
            )
            ws.cell(
                row=annotation_row, column=2,
                value=(
                    f"{data[best_hour_start].get('interval', '')} - "
                    f"{data[best_hour_start + 3].get('interval', '')} — "
                    f"{best_hour_vol} vehicles"
                ),
            )
