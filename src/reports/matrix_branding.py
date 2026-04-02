"""Matrix Traffic Data — Report branding constants and helper functions.

Provides consistent Matrix-branded styling for all Excel reports generated
by the reporting engine. Uses openpyxl styles throughout.
"""

from __future__ import annotations

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.common.data_models import JobConfig

# ── Color constants (hex without leading #) ──

HEADER_BG: str = "1B2A4A"       # Matrix navy (from logo)
HEADER_TEXT: str = "FFFFFF"      # White
SUBHEADER_BG: str = "E9EDF3"    # Light navy tint
ACCENT_BLUE: str = "1B2A4A"     # Matrix navy (brand primary)
TOTAL_ROW_BG: str = "F7F8FA"    # Very light gray
BORDER_COLOR: str = "D1D5DB"    # Medium gray
ALT_ROW_BG: str = "FAFBFC"      # Very subtle alternate row


# ── Style helper functions ──

def get_header_font() -> Font:
    """Return the standard Matrix header font (white, bold, 11pt)."""
    return Font(name="Calibri", bold=True, size=11, color=HEADER_TEXT)


def get_header_fill() -> PatternFill:
    """Return the standard Matrix header background fill (dark navy)."""
    return PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")


def get_border() -> Border:
    """Return a thin border using the Matrix border color."""
    side = Side(style="thin", color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def get_total_fill() -> PatternFill:
    """Return the fill used for total / summary rows."""
    return PatternFill(start_color=TOTAL_ROW_BG, end_color=TOTAL_ROW_BG, fill_type="solid")


def get_subheader_fill() -> PatternFill:
    """Return the fill used for sub-header rows."""
    return PatternFill(start_color=SUBHEADER_BG, end_color=SUBHEADER_BG, fill_type="solid")


def get_alt_row_fill() -> PatternFill:
    """Return the fill used for alternating data rows."""
    return PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")


def get_accent_font() -> Font:
    """Return a font styled with the brand accent blue."""
    return Font(name="Calibri", bold=True, size=11, color=ACCENT_BLUE)


# ── Header writer ──

def write_matrix_header(
    ws: Worksheet,
    job_config: JobConfig,
    report_title: str,
    start_row: int = 1,
) -> int:
    """Write the standard Matrix report header block to a worksheet.

    Layout
    ------
    Row 1: "MATRIX TRAFFIC DATA" — large bold navy
    Row 2: Report type subtitle (e.g. "Automatic Traffic Count Report")
    Row 3: blank
    Row 4: "Job: {number} - {name}"
    Row 5: "Site: {site_name}"
    Row 6: "Survey Date: {date or date range}"
    Row 7: "Survey Period: {start_time} - {end_time}"
    Row 8: blank separator

    Parameters
    ----------
    ws : Worksheet
        Target worksheet.
    job_config : JobConfig
        Job configuration with survey metadata.
    report_title : str
        Report type subtitle (row 2).
    start_row : int, optional
        First row to begin writing (default 1).

    Returns
    -------
    int
        The next available row number after the header block.
    """
    row = start_row

    # Row 1 — company title
    title_font = Font(name="Calibri", bold=True, size=16, color=HEADER_BG)
    ws.cell(row=row, column=1, value="MATRIX TRAFFIC DATA").font = title_font
    row += 1

    # Row 2 — report subtitle
    subtitle_font = Font(name="Calibri", bold=True, size=12, color=ACCENT_BLUE)
    ws.cell(row=row, column=1, value=report_title).font = subtitle_font
    row += 1

    # Row 3 — blank
    row += 1

    # Row 4 — job info
    label_font = Font(name="Calibri", bold=True, size=10, color=HEADER_BG)
    value_font = Font(name="Calibri", size=10)

    ws.cell(row=row, column=1, value="Job:").font = label_font
    ws.cell(row=row, column=2, value=job_config.display_title).font = value_font
    row += 1

    # Row 5 — site name (first site or summary)
    site_name = ""
    if job_config.sites:
        site_name = job_config.sites[0].display_name
    ws.cell(row=row, column=1, value="Site:").font = label_font
    ws.cell(row=row, column=2, value=site_name).font = value_font
    row += 1

    # Row 6 — survey date
    ws.cell(row=row, column=1, value="Survey Date:").font = label_font
    ws.cell(row=row, column=2, value=job_config.date_display).font = value_font
    row += 1

    # Row 7 — survey period
    ws.cell(row=row, column=1, value="Survey Period:").font = label_font
    ws.cell(row=row, column=2, value=job_config.time_display).font = value_font
    row += 1

    # Row 8 — blank separator
    row += 1

    return row


# ── Row / range styling helpers ──

def apply_header_style(
    ws: Worksheet,
    row: int,
    col_start: int,
    col_end: int,
) -> None:
    """Apply Matrix header formatting (navy bg, white bold text) to a row range.

    Parameters
    ----------
    ws : Worksheet
        Target worksheet.
    row : int
        The row number to style.
    col_start : int
        First column (1-based).
    col_end : int
        Last column (1-based, inclusive).
    """
    hdr_font = get_header_font()
    hdr_fill = get_header_fill()
    hdr_border = get_border()
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = hdr_border
        cell.alignment = hdr_align


def apply_data_borders(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    col_start: int,
    col_end: int,
) -> None:
    """Apply thin borders and alternating row shading to a data range.

    Parameters
    ----------
    ws : Worksheet
        Target worksheet.
    start_row : int
        First data row (1-based).
    end_row : int
        Last data row (1-based, inclusive).
    col_start : int
        First column (1-based).
    col_end : int
        Last column (1-based, inclusive).
    """
    border = get_border()
    alt_fill = get_alt_row_fill()
    data_align = Alignment(horizontal="center", vertical="center")

    for r in range(start_row, end_row + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = data_align
        # Alternate row shading (even rows relative to start)
        if (r - start_row) % 2 == 1:
            for c in range(col_start, col_end + 1):
                ws.cell(row=r, column=c).fill = alt_fill


def auto_size_columns(ws: Worksheet, min_width: float = 8.0, max_width: float = 40.0) -> None:
    """Auto-fit column widths based on cell content.

    Parameters
    ----------
    ws : Worksheet
        Target worksheet.
    min_width : float
        Minimum column width in characters.
    max_width : float
        Maximum column width in characters.
    """
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted
