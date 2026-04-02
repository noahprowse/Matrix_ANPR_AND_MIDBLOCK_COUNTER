"""Matrix Traffic Data — Midblock Automatic Traffic Count (ATC) Report.

Generates a branded Excel workbook with Summary, 15-Min Data, Hourly Summary,
and Classification sheets for midblock counting jobs.
"""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.common.data_models import AUSTROADS_CLASSES, JobConfig, SiteConfig
from src.reports.matrix_branding import (
    ACCENT_BLUE,
    HEADER_BG,
    apply_data_borders,
    apply_header_style,
    auto_size_columns,
    get_border,
    get_header_fill,
    get_header_font,
    get_total_fill,
)
from src.reports.report_engine import BaseReport


class ATCReport(BaseReport):
    """Midblock Automatic Traffic Count report.

    Expected ``results`` structure passed to ``generate()``::

        {
            "directions": ["Northbound", "Southbound"],
            "class_codes": ["1", "1M", "3", ...],
            "totals": {
                "Northbound": {"1": 1200, "1M": 45, "3": 80, ...},
                "Southbound": {"1": 1150, "1M": 38, "3": 72, ...},
            },
            "interval_data": [
                {
                    "interval": "07:00",
                    "counts": {
                        "Northbound": {"1": 25, "1M": 1, ...},
                        "Southbound": {"1": 22, "1M": 0, ...},
                    },
                },
                ...
            ],
            "hourly_data": [
                {
                    "interval": "07:00",
                    "counts": { ... same structure ... },
                },
                ...
            ],
        }
    """

    def __init__(self, job_config: JobConfig) -> None:
        super().__init__(job_config)

    def generate(self, results: dict[str, Any], site_config: SiteConfig) -> None:
        """Build the full ATC report workbook.

        Parameters
        ----------
        results : dict
            Counting results (see class docstring for structure).
        site_config : SiteConfig
            Site-level configuration for display labels.
        """
        wb = self._create_workbook()

        directions: list[str] = results.get("directions", [])
        class_codes: list[str] = results.get("class_codes", [])
        totals: dict[str, dict[str, int]] = results.get("totals", {})
        interval_data: list[dict[str, Any]] = results.get("interval_data", [])
        hourly_data: list[dict[str, Any]] = results.get("hourly_data", [])

        self._build_summary_sheet(wb, directions, class_codes, totals, site_config)
        self._build_15min_sheet(wb, interval_data, class_codes, directions)
        self._build_hourly_sheet(wb, hourly_data, class_codes, directions)
        self._build_classification_sheet(wb, class_codes, totals, directions)

    # ── Sheet builders ──

    def _build_summary_sheet(
        self,
        wb: Any,
        directions: list[str],
        class_codes: list[str],
        totals: dict[str, dict[str, int]],
        site_config: SiteConfig,
    ) -> None:
        """Sheet 'Summary': totals by class and direction (In/Out/Total)."""
        ws = self._add_sheet(wb, "Summary")
        row = self._write_header(ws, "Automatic Traffic Count Report")

        # Build header: Class Name | Dir1 | Dir2 | ... | Total
        headers = ["Class"]
        for d in directions:
            headers.append(d)
        headers.append("Total")

        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=row, column=col_idx, value=h)
        apply_header_style(ws, row, 1, len(headers))

        data_start = row + 1
        current_row = data_start

        for cc in class_codes:
            class_info = AUSTROADS_CLASSES.get(cc, {})
            class_name = class_info.get("name", cc)
            ws.cell(row=current_row, column=1, value=f"{cc} - {class_name}")
            row_total = 0
            for col_idx, d in enumerate(directions, start=2):
                val = totals.get(d, {}).get(cc, 0)
                ws.cell(row=current_row, column=col_idx, value=val)
                row_total += val
            ws.cell(row=current_row, column=len(headers), value=row_total)
            current_row += 1

        data_end = current_row - 1
        apply_data_borders(ws, data_start, data_end, 1, len(headers))

        # Total row
        total_fill = get_total_fill()
        border = get_border()
        ws.cell(row=current_row, column=1, value="TOTAL")
        ws.cell(row=current_row, column=1).font = Font(name="Calibri", bold=True, size=11)

        for c in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=c).fill = total_fill
            ws.cell(row=current_row, column=c).border = border
            ws.cell(row=current_row, column=c).font = Font(name="Calibri", bold=True, size=11)
            ws.cell(row=current_row, column=c).alignment = Alignment(horizontal="center")

        for col_idx, d in enumerate(directions, start=2):
            dir_total = sum(totals.get(d, {}).get(cc, 0) for cc in class_codes)
            ws.cell(row=current_row, column=col_idx, value=dir_total)

        grand_total = sum(
            sum(totals.get(d, {}).get(cc, 0) for cc in class_codes)
            for d in directions
        )
        ws.cell(row=current_row, column=len(headers), value=grand_total)

        auto_size_columns(ws)

    def _build_15min_sheet(
        self,
        wb: Any,
        interval_data: list[dict[str, Any]],
        class_codes: list[str],
        directions: list[str],
    ) -> None:
        """Sheet '15-Min Data': per 15-min interval counts with peak highlight."""
        ws = self._add_sheet(wb, "15-Min Data")
        row = self._write_header(ws, "15-Minute Interval Data")

        next_row = self._write_interval_table(ws, interval_data, class_codes, row, directions)

        # Peak identification
        if interval_data:
            peak_start, peak_end, peak_vol = self._find_peak_hour(interval_data)

            # Determine column count for highlighting
            col_count = 1 + len(directions) * (len(class_codes) + 1) + 1

            # Highlight peak 15-min (single highest interval)
            best_single_idx = 0
            best_single_vol = 0
            for i, rec in enumerate(interval_data):
                rec_total = sum(
                    sum(dc.values()) for dc in rec.get("counts", {}).values()
                )
                if rec_total > best_single_vol:
                    best_single_vol = rec_total
                    best_single_idx = i

            # data starts at row + 1 (after header row written by _write_interval_table)
            data_row_start = row + 1
            self._highlight_peak_row(ws, data_row_start + best_single_idx, 1, col_count)

            # Highlight peak hour (4 consecutive intervals)
            for i in range(peak_start, peak_end + 1):
                peak_row = data_row_start + i
                peak_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                peak_font = Font(name="Calibri", bold=True, size=11, color=ACCENT_BLUE)
                for c in range(1, col_count + 1):
                    cell = ws.cell(row=peak_row, column=c)
                    # Don't overwrite the single-peak highlight
                    if i != best_single_idx:
                        cell.fill = peak_fill
                        cell.font = peak_font

            # Peak hour annotation
            annotation_row = next_row + 1
            ws.cell(row=annotation_row, column=1, value="Peak 15-Min Interval:").font = Font(
                name="Calibri", bold=True, size=10, color=HEADER_BG
            )
            ws.cell(
                row=annotation_row, column=2,
                value=f"{interval_data[best_single_idx].get('interval', '')} — {best_single_vol} vehicles",
            )
            annotation_row += 1
            ws.cell(row=annotation_row, column=1, value="Peak Hour:").font = Font(
                name="Calibri", bold=True, size=10, color=HEADER_BG
            )
            peak_start_label = interval_data[peak_start].get("interval", "")
            peak_end_label = interval_data[peak_end].get("interval", "")
            ws.cell(
                row=annotation_row, column=2,
                value=f"{peak_start_label} - {peak_end_label} — {peak_vol} vehicles",
            )

        auto_size_columns(ws)

    def _build_hourly_sheet(
        self,
        wb: Any,
        hourly_data: list[dict[str, Any]],
        class_codes: list[str],
        directions: list[str],
    ) -> None:
        """Sheet 'Hourly Summary': hourly aggregation with peak hour."""
        ws = self._add_sheet(wb, "Hourly Summary")
        row = self._write_header(ws, "Hourly Summary")

        next_row = self._write_interval_table(ws, hourly_data, class_codes, row, directions)

        # Peak hour = single highest hourly row
        if hourly_data:
            best_idx = 0
            best_vol = 0
            for i, rec in enumerate(hourly_data):
                rec_total = sum(
                    sum(dc.values()) for dc in rec.get("counts", {}).values()
                )
                if rec_total > best_vol:
                    best_vol = rec_total
                    best_idx = i

            col_count = 1 + len(directions) * (len(class_codes) + 1) + 1
            data_row_start = row + 1
            self._highlight_peak_row(ws, data_row_start + best_idx, 1, col_count)

            annotation_row = next_row + 1
            ws.cell(row=annotation_row, column=1, value="Peak Hour:").font = Font(
                name="Calibri", bold=True, size=10, color=HEADER_BG
            )
            ws.cell(
                row=annotation_row, column=2,
                value=f"{hourly_data[best_idx].get('interval', '')} — {best_vol} vehicles",
            )

        auto_size_columns(ws)

    def _build_classification_sheet(
        self,
        wb: Any,
        class_codes: list[str],
        totals: dict[str, dict[str, int]],
        directions: list[str],
    ) -> None:
        """Sheet 'Classification': class breakdown percentages."""
        ws = self._add_sheet(wb, "Classification")
        row = self._write_header(ws, "Vehicle Classification Summary")

        # Compute grand totals per direction and overall
        dir_grand: dict[str, int] = {}
        for d in directions:
            dir_grand[d] = sum(totals.get(d, {}).get(cc, 0) for cc in class_codes)
        overall_grand = sum(dir_grand.values())

        # Headers: Class | Dir1 Count | Dir1 % | Dir2 Count | Dir2 % | Total Count | Total %
        headers: list[str] = ["Class"]
        for d in directions:
            headers.append(f"{d} Count")
            headers.append(f"{d} %")
        headers.append("Total Count")
        headers.append("Total %")

        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=row, column=col_idx, value=h)
        apply_header_style(ws, row, 1, len(headers))

        data_start = row + 1
        current_row = data_start

        for cc in class_codes:
            class_info = AUSTROADS_CLASSES.get(cc, {})
            class_name = class_info.get("name", cc)
            ws.cell(row=current_row, column=1, value=f"{cc} - {class_name}")

            col = 2
            row_total = 0
            for d in directions:
                count = totals.get(d, {}).get(cc, 0)
                row_total += count
                ws.cell(row=current_row, column=col, value=count)
                col += 1
                pct = (count / dir_grand[d] * 100) if dir_grand[d] > 0 else 0.0
                cell = ws.cell(row=current_row, column=col, value=round(pct, 1))
                cell.number_format = "0.0"
                col += 1

            ws.cell(row=current_row, column=col, value=row_total)
            col += 1
            overall_pct = (row_total / overall_grand * 100) if overall_grand > 0 else 0.0
            cell = ws.cell(row=current_row, column=col, value=round(overall_pct, 1))
            cell.number_format = "0.0"

            current_row += 1

        data_end = current_row - 1
        apply_data_borders(ws, data_start, data_end, 1, len(headers))

        # Total row
        total_fill = get_total_fill()
        border = get_border()
        ws.cell(row=current_row, column=1, value="TOTAL")

        col = 2
        for d in directions:
            ws.cell(row=current_row, column=col, value=dir_grand[d])
            col += 1
            ws.cell(row=current_row, column=col, value=100.0)
            ws.cell(row=current_row, column=col).number_format = "0.0"
            col += 1
        ws.cell(row=current_row, column=col, value=overall_grand)
        col += 1
        ws.cell(row=current_row, column=col, value=100.0)
        ws.cell(row=current_row, column=col).number_format = "0.0"

        for c in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=c).fill = total_fill
            ws.cell(row=current_row, column=c).border = border
            ws.cell(row=current_row, column=c).font = Font(name="Calibri", bold=True, size=11)
            ws.cell(row=current_row, column=c).alignment = Alignment(horizontal="center")

        auto_size_columns(ws)
