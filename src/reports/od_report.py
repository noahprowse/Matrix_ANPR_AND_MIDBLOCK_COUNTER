"""Matrix Traffic Data — Origin-Destination (O-D) Report.

Generates a branded Excel workbook with O-D Total, O-D 15-Min, and O-D by Class
sheets for intersection or zone-based O-D analysis.
"""

from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from src.common.data_models import AUSTROADS_CLASSES, JobConfig, SiteConfig
from src.reports.matrix_branding import (
    ACCENT_BLUE,
    HEADER_BG,
    apply_data_borders,
    apply_header_style,
    auto_size_columns,
    get_border,
    get_total_fill,
)
from src.reports.report_engine import BaseReport


class ODReport(BaseReport):
    """Origin-Destination report.

    Expected ``od_matrix`` structure passed to ``generate()``::

        {
            "origins": ["North", "South", "East", "West"],
            "destinations": ["North", "South", "East", "West"],
            "matrix": {
                "North": {"North": 0, "South": 120, "East": 80, "West": 50},
                "South": {"North": 115, "South": 0, "East": 65, "West": 42},
                ...
            },
            "by_class": {
                "1": {
                    "North": {"North": 0, "South": 100, ...},
                    ...
                },
                "3": { ... },
                ...
            },
            "interval_matrices": [
                {
                    "interval": "07:00",
                    "matrix": {
                        "North": {"North": 0, "South": 12, ...},
                        ...
                    },
                },
                ...
            ],
            "class_codes": ["1", "1M", "3", ...],
        }
    """

    def __init__(self, job_config: JobConfig) -> None:
        super().__init__(job_config)

    def generate(self, od_matrix: dict[str, Any], site_config: SiteConfig) -> None:
        """Build the full O-D report workbook.

        Parameters
        ----------
        od_matrix : dict
            O-D matrix data (see class docstring for structure).
        site_config : SiteConfig
            Site-level configuration for display labels.
        """
        wb = self._create_workbook()

        origins: list[str] = od_matrix.get("origins", [])
        destinations: list[str] = od_matrix.get("destinations", [])
        matrix: dict[str, dict[str, int]] = od_matrix.get("matrix", {})
        by_class: dict[str, dict[str, dict[str, int]]] = od_matrix.get("by_class", {})
        interval_matrices: list[dict[str, Any]] = od_matrix.get("interval_matrices", [])
        class_codes: list[str] = od_matrix.get("class_codes", [])

        self._build_total_sheet(wb, origins, destinations, matrix)
        self._build_15min_sheet(wb, origins, destinations, interval_matrices)
        self._build_by_class_sheet(wb, origins, destinations, by_class, class_codes)

    # ── Sheet builders ──

    def _build_total_sheet(
        self,
        wb: Workbook,
        origins: list[str],
        destinations: list[str],
        matrix: dict[str, dict[str, int]],
    ) -> None:
        """Sheet 'O-D Total': complete O-D matrix across all intervals."""
        ws = self._add_sheet(wb, "O-D Total")
        row = self._write_header(ws, "Origin-Destination Matrix — Total")

        row = self._write_od_matrix(ws, origins, destinations, matrix, row)

        # Match percentages
        row += 1
        self._write_match_percentages(ws, origins, destinations, matrix, row)

        auto_size_columns(ws)

    def _build_15min_sheet(
        self,
        wb: Workbook,
        origins: list[str],
        destinations: list[str],
        interval_matrices: list[dict[str, Any]],
    ) -> None:
        """Sheet 'O-D 15-Min': one matrix per 15-min interval, stacked vertically."""
        ws = self._add_sheet(wb, "O-D 15-Min")
        row = self._write_header(ws, "Origin-Destination Matrix — 15-Minute Intervals")

        if not interval_matrices:
            ws.cell(row=row, column=1, value="No interval O-D data available.")
            auto_size_columns(ws)
            return

        for record in interval_matrices:
            interval_label = record.get("interval", "")
            interval_matrix = record.get("matrix", {})

            # Interval label
            ws.cell(row=row, column=1, value=f"Interval: {interval_label}")
            ws.cell(row=row, column=1).font = Font(
                name="Calibri", bold=True, size=11, color=ACCENT_BLUE
            )
            row += 1

            row = self._write_od_matrix(ws, origins, destinations, interval_matrix, row)
            row += 1  # Blank separator between intervals

        auto_size_columns(ws)

    def _build_by_class_sheet(
        self,
        wb: Workbook,
        origins: list[str],
        destinations: list[str],
        by_class: dict[str, dict[str, dict[str, int]]],
        class_codes: list[str],
    ) -> None:
        """Sheet 'O-D by Class': separate O-D matrix per vehicle class."""
        ws = self._add_sheet(wb, "O-D by Class")
        row = self._write_header(ws, "Origin-Destination Matrix — By Vehicle Class")

        if not by_class:
            ws.cell(row=row, column=1, value="No class-level O-D data available.")
            auto_size_columns(ws)
            return

        for cc in class_codes:
            class_matrix = by_class.get(cc)
            if class_matrix is None:
                continue

            class_info = AUSTROADS_CLASSES.get(cc, {})
            class_name = class_info.get("name", cc)

            # Class label
            ws.cell(row=row, column=1, value=f"Class {cc} — {class_name}")
            ws.cell(row=row, column=1).font = Font(
                name="Calibri", bold=True, size=11, color=ACCENT_BLUE
            )
            row += 1

            row = self._write_od_matrix(ws, origins, destinations, class_matrix, row)
            row += 1  # Blank separator between classes

        auto_size_columns(ws)

    # ── Reusable O-D matrix writer ──

    def _write_od_matrix(
        self,
        ws: Worksheet,
        origins: list[str],
        destinations: list[str],
        matrix: dict[str, dict[str, int]],
        start_row: int,
    ) -> int:
        """Write a single O-D matrix table with total row and column.

        Parameters
        ----------
        ws : Worksheet
            Target worksheet.
        origins : list[str]
            Row labels (origin zones).
        destinations : list[str]
            Column labels (destination zones).
        matrix : dict[str, dict[str, int]]
            Nested dict: matrix[origin][destination] = count.
        start_row : int
            Row to begin writing.

        Returns
        -------
        int
            Next available row after the matrix (including total row).
        """
        if not origins or not destinations:
            ws.cell(row=start_row, column=1, value="No O-D data available.")
            return start_row + 1

        # Header row: "Origin \\ Dest" | Dest1 | Dest2 | ... | Total
        headers = ["Origin \\ Dest"]
        headers.extend(destinations)
        headers.append("Total")
        col_end = len(headers)

        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=start_row, column=col_idx, value=h)
        apply_header_style(ws, start_row, 1, col_end)

        # Data rows
        data_start = start_row + 1
        current_row = data_start
        total_fill = get_total_fill()
        border = get_border()

        col_totals: dict[str, int] = {d: 0 for d in destinations}
        grand_total = 0

        for origin in origins:
            ws.cell(row=current_row, column=1, value=origin)
            row_total = 0
            for col_idx, dest in enumerate(destinations, start=2):
                val = matrix.get(origin, {}).get(dest, 0)
                ws.cell(row=current_row, column=col_idx, value=val)
                row_total += val
                col_totals[dest] += val
            ws.cell(row=current_row, column=col_end, value=row_total)
            grand_total += row_total
            current_row += 1

        data_end = current_row - 1
        if origins:
            apply_data_borders(ws, data_start, data_end, 1, col_end)

        # Total row
        ws.cell(row=current_row, column=1, value="Total")
        for col_idx, dest in enumerate(destinations, start=2):
            ws.cell(row=current_row, column=col_idx, value=col_totals[dest])
        ws.cell(row=current_row, column=col_end, value=grand_total)

        for c in range(1, col_end + 1):
            ws.cell(row=current_row, column=c).fill = total_fill
            ws.cell(row=current_row, column=c).border = border
            ws.cell(row=current_row, column=c).font = Font(name="Calibri", bold=True, size=11)
            ws.cell(row=current_row, column=c).alignment = Alignment(horizontal="center")

        return current_row + 1

    # ── Match percentage calculation ──

    def _write_match_percentages(
        self,
        ws: Worksheet,
        origins: list[str],
        destinations: list[str],
        matrix: dict[str, dict[str, int]],
        start_row: int,
    ) -> int:
        """Write an O-D match-percentage matrix showing row-wise distribution.

        Each cell shows what percentage of the origin's total traffic goes to
        each destination.

        Parameters
        ----------
        ws : Worksheet
            Target worksheet.
        origins : list[str]
            Origin zone labels.
        destinations : list[str]
            Destination zone labels.
        matrix : dict[str, dict[str, int]]
            The O-D count matrix.
        start_row : int
            Row to begin writing.

        Returns
        -------
        int
            Next available row after the percentage matrix.
        """
        # Section label
        ws.cell(row=start_row, column=1, value="Match Percentages (Row Distribution)")
        ws.cell(row=start_row, column=1).font = Font(
            name="Calibri", bold=True, size=11, color=HEADER_BG
        )
        start_row += 1

        # Header row
        headers = ["Origin \\ Dest"]
        headers.extend(destinations)
        headers.append("Total")
        col_end = len(headers)

        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=start_row, column=col_idx, value=h)
        apply_header_style(ws, start_row, 1, col_end)

        data_start = start_row + 1
        current_row = data_start

        for origin in origins:
            ws.cell(row=current_row, column=1, value=origin)
            row_total = sum(matrix.get(origin, {}).get(d, 0) for d in destinations)

            for col_idx, dest in enumerate(destinations, start=2):
                val = matrix.get(origin, {}).get(dest, 0)
                pct = (val / row_total * 100) if row_total > 0 else 0.0
                cell = ws.cell(row=current_row, column=col_idx, value=round(pct, 1))
                cell.number_format = "0.0"

            ws.cell(row=current_row, column=col_end, value=100.0 if row_total > 0 else 0.0)
            ws.cell(row=current_row, column=col_end).number_format = "0.0"
            current_row += 1

        data_end = current_row - 1
        if origins:
            apply_data_borders(ws, data_start, data_end, 1, col_end)

        return current_row
