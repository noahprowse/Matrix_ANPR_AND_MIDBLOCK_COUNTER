"""Matrix Traffic Data — Base report builder.

Provides the ``BaseReport`` class with common workbook creation, header
writing, data table rendering, and peak-hour identification used by all
concrete report types (ATC, TMC, O-D).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.common.data_models import JobConfig
from src.reports.matrix_branding import (
    ACCENT_BLUE,
    HEADER_BG,
    TOTAL_ROW_BG,
    apply_data_borders,
    apply_header_style,
    auto_size_columns,
    get_border,
    get_header_fill,
    get_header_font,
    get_total_fill,
    write_matrix_header,
)


class BaseReport:
    """Base class for all Matrix Excel reports.

    Subclasses implement a ``generate()`` method that populates the workbook
    with domain-specific sheets, then call ``save()`` to write to disk.
    """

    def __init__(self, job_config: JobConfig) -> None:
        self.job_config: JobConfig = job_config
        self.wb: Optional[Workbook] = None

    # ── Workbook lifecycle ──

    def _create_workbook(self) -> Workbook:
        """Create a new openpyxl Workbook and store it on the instance."""
        self.wb = Workbook()
        # Remove the default sheet — we will add named ones explicitly
        default_sheet = self.wb.active
        if default_sheet is not None:
            self.wb.remove(default_sheet)
        return self.wb

    def _add_sheet(self, wb: Workbook, title: str) -> Worksheet:
        """Add a new worksheet with the given title.

        Parameters
        ----------
        wb : Workbook
            The target workbook.
        title : str
            Sheet tab name (truncated to 31 chars by openpyxl).

        Returns
        -------
        Worksheet
        """
        return wb.create_sheet(title=title[:31])

    # ── Header ──

    def _write_header(self, ws: Worksheet, title: str) -> int:
        """Write the Matrix branded header using ``write_matrix_header``.

        Returns the next available row after the header block.
        """
        return write_matrix_header(ws, self.job_config, title)

    # ── Generic data table ──

    def _write_data_table(
        self,
        ws: Worksheet,
        headers: list[str],
        data: list[list[Any]],
        start_row: int,
    ) -> int:
        """Write a generic table with headers and data rows.

        Parameters
        ----------
        ws : Worksheet
            Target worksheet.
        headers : list[str]
            Column header labels.
        data : list[list[Any]]
            Row data — each inner list corresponds to one row.
        start_row : int
            Row number for the header row.

        Returns
        -------
        int
            The next available row after the table.
        """
        col_end = len(headers)

        # Header row
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=start_row, column=col_idx, value=header)
        apply_header_style(ws, start_row, 1, col_end)

        # Data rows
        data_start = start_row + 1
        for row_offset, row_data in enumerate(data):
            row_num = data_start + row_offset
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_num, column=col_idx, value=value)

        data_end = data_start + len(data) - 1
        if data:
            apply_data_borders(ws, data_start, data_end, 1, col_end)

        return data_end + 1 if data else data_start

    # ── Interval table ──

    def _write_interval_table(
        self,
        ws: Worksheet,
        interval_data: list[dict[str, Any]],
        class_codes: list[str],
        start_row: int,
        directions: Optional[list[str]] = None,
    ) -> int:
        """Write a 15-min (or hourly) interval table with class columns per direction.

        Expected ``interval_data`` structure::

            [
                {
                    "interval": "07:00",
                    "counts": {
                        "Direction1": {"1": 5, "3": 2, ...},
                        "Direction2": {"1": 8, "3": 1, ...},
                    },
                },
                ...
            ]

        Parameters
        ----------
        ws : Worksheet
            Target worksheet.
        interval_data : list[dict]
            Interval records (see structure above).
        class_codes : list[str]
            Ordered Austroads class codes for columns.
        start_row : int
            Row to begin writing.
        directions : list[str] | None
            Direction labels. If None, derived from first interval record.

        Returns
        -------
        int
            Next available row after the table.
        """
        if not interval_data:
            return start_row

        # Resolve direction labels
        if directions is None:
            first_counts = interval_data[0].get("counts", {})
            directions = list(first_counts.keys())

        # Build header row: Time | Dir1-Class1 | Dir1-Class2 | ... | Total Dir1 | Dir2-... | Total Dir2 | Grand Total
        headers: list[str] = ["Time"]
        for d in directions:
            for cc in class_codes:
                headers.append(f"{d} {cc}")
            headers.append(f"Total {d}")
        headers.append("Grand Total")

        col_end = len(headers)

        # Write header
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=start_row, column=col_idx, value=h)
        apply_header_style(ws, start_row, 1, col_end)

        # Write data
        data_start = start_row + 1
        for row_offset, record in enumerate(interval_data):
            row_num = data_start + row_offset
            col = 1
            ws.cell(row=row_num, column=col, value=record.get("interval", ""))
            col += 1
            grand_total = 0
            for d in directions:
                dir_counts = record.get("counts", {}).get(d, {})
                dir_total = 0
                for cc in class_codes:
                    val = dir_counts.get(cc, 0)
                    ws.cell(row=row_num, column=col, value=val)
                    dir_total += val
                    col += 1
                ws.cell(row=row_num, column=col, value=dir_total)
                grand_total += dir_total
                col += 1
            ws.cell(row=row_num, column=col, value=grand_total)

        data_end = data_start + len(interval_data) - 1
        apply_data_borders(ws, data_start, data_end, 1, col_end)

        # Total row
        total_row = data_end + 1
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=1).font = Font(name="Calibri", bold=True, size=11)
        total_fill = get_total_fill()
        border = get_border()

        for c in range(1, col_end + 1):
            col_sum = 0
            for r in range(data_start, data_end + 1):
                cell_val = ws.cell(row=r, column=c).value
                if isinstance(cell_val, (int, float)):
                    col_sum += cell_val
            if c > 1:
                ws.cell(row=total_row, column=c, value=col_sum)
            ws.cell(row=total_row, column=c).fill = total_fill
            ws.cell(row=total_row, column=c).border = border
            ws.cell(row=total_row, column=c).font = Font(name="Calibri", bold=True, size=11)
            ws.cell(row=total_row, column=c).alignment = Alignment(horizontal="center")

        return total_row + 1

    # ── Peak-hour identification ──

    def _find_peak_hour(
        self,
        interval_data: list[dict[str, Any]],
    ) -> tuple[int, int, int]:
        """Find the consecutive 4x 15-min intervals with the highest total.

        Parameters
        ----------
        interval_data : list[dict]
            Same structure as ``_write_interval_table``.

        Returns
        -------
        tuple[int, int, int]
            (start_index, end_index, peak_volume) where indices are 0-based
            into ``interval_data``. ``end_index`` is inclusive.
            Returns (0, 0, 0) if fewer than 4 intervals exist.
        """
        if len(interval_data) < 4:
            return (0, 0, 0)

        def _interval_total(record: dict) -> int:
            total = 0
            for dir_counts in record.get("counts", {}).values():
                total += sum(dir_counts.values())
            return total

        best_start = 0
        best_total = 0
        for i in range(len(interval_data) - 3):
            window_total = sum(_interval_total(interval_data[j]) for j in range(i, i + 4))
            if window_total > best_total:
                best_total = window_total
                best_start = i

        return (best_start, best_start + 3, best_total)

    def _highlight_peak_row(
        self,
        ws: Worksheet,
        row: int,
        col_start: int,
        col_end: int,
    ) -> None:
        """Highlight a row with the accent-blue fill to mark peak intervals.

        Parameters
        ----------
        ws : Worksheet
            Target worksheet.
        row : int
            Row number to highlight.
        col_start : int
            First column (1-based).
        col_end : int
            Last column (1-based, inclusive).
        """
        peak_fill = PatternFill(start_color=ACCENT_BLUE, end_color=ACCENT_BLUE, fill_type="solid")
        peak_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")

        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = peak_fill
            cell.font = peak_font

    # ── Persistence ──

    def save(self, filepath: str | Path) -> Path:
        """Save the workbook to disk.

        Parameters
        ----------
        filepath : str | Path
            Destination file path (should end with ``.xlsx``).

        Returns
        -------
        Path
            The resolved path the file was saved to.

        Raises
        ------
        RuntimeError
            If no workbook has been created yet.
        """
        if self.wb is None:
            raise RuntimeError("No workbook to save. Call generate() first.")

        dest = Path(filepath)
        dest.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(dest))
        return dest
