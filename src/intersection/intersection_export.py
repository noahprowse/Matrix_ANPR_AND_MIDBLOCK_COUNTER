"""Excel export for Intersection TMC and O-D matrix results.

Uses Matrix branding colors and openpyxl for formatted Excel output.
"""

from __future__ import annotations

import logging
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.common.data_models import AUSTROADS_CLASSES, JobConfig
from src.intersection.od_matrix import ODMatrix
from src.intersection.tmc_calculator import MOVEMENT_ORDER, TMCCalculator

logger = logging.getLogger(__name__)

# ── Matrix branding colors ────────────────────────────────────────────

_HEADER_NAVY = "1A365D"
_ACCENT_BLUE = "2563EB"
_LIGHT_GRAY = "E2E8F0"
_WHITE = "FFFFFF"
_LIGHT_BLUE_BG = "EFF6FF"

_header_fill = PatternFill(
    start_color=_HEADER_NAVY, end_color=_HEADER_NAVY, fill_type="solid"
)
_header_font = Font(color=_WHITE, bold=True, size=11, name="Segoe UI")
_accent_fill = PatternFill(
    start_color=_ACCENT_BLUE, end_color=_ACCENT_BLUE, fill_type="solid"
)
_accent_font = Font(color=_WHITE, bold=True, size=11, name="Segoe UI")
_light_fill = PatternFill(
    start_color=_LIGHT_BLUE_BG, end_color=_LIGHT_BLUE_BG, fill_type="solid"
)
_thin_border = Border(
    left=Side(style="thin", color=_LIGHT_GRAY),
    right=Side(style="thin", color=_LIGHT_GRAY),
    top=Side(style="thin", color=_LIGHT_GRAY),
    bottom=Side(style="thin", color=_LIGHT_GRAY),
)
_center_align = Alignment(horizontal="center", vertical="center")
_left_align = Alignment(horizontal="left", vertical="center")


class IntersectionExporter:
    """Exports intersection TMC and O-D results to formatted Excel."""

    def export_tmc(
        self,
        filepath: str,
        job_config: JobConfig,
        tmc_data: dict,
        od_matrix: ODMatrix,
    ) -> None:
        """Write a multi-sheet Excel workbook.

        Sheets:
            - TMC Summary: approach x movement x class totals
            - TMC 15-min: per-interval TMC
            - O-D Matrix: full O-D matrix (all classes summed)
            - O-D by Class: one section per vehicle class

        Args:
            filepath:   Output .xlsx path.
            job_config: Job configuration for report header.
            tmc_data:   TMC data from TMCCalculator.compute_tmc().
            od_matrix:  The ODMatrix instance with accumulated data.
        """
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            self._write_tmc_summary(writer, tmc_data, od_matrix, job_config)
            self._write_tmc_interval(writer, tmc_data, od_matrix)
            self._write_od_matrix(writer, od_matrix)
            self._write_od_by_class(writer, od_matrix)

        self._format_workbook(filepath, job_config, od_matrix)
        logger.info("Intersection TMC exported to %s", filepath)

    # ── Sheet writers ─────────────────────────────────────────────────

    def _write_tmc_summary(
        self,
        writer: pd.ExcelWriter,
        tmc_data: dict,
        od_matrix: ODMatrix,
        job_config: JobConfig,
    ) -> None:
        """TMC Summary sheet: approach x movement x class totals."""
        # Build approach config from zones in tmc_data
        # Compute total TMC across all intervals
        total_tmc = self._compute_total_tmc(tmc_data)

        class_codes = od_matrix.get_class_codes() or ["1"]
        rows = []

        for approach, movements in sorted(total_tmc.items()):
            for movement in MOVEMENT_ORDER:
                classes = movements.get(movement, {})
                row = {
                    "Approach": approach,
                    "Movement": movement,
                }
                movement_total = 0
                for cls in class_codes:
                    cls_name = AUSTROADS_CLASSES.get(cls, {}).get("name", cls)
                    count = classes.get(cls, 0)
                    row[f"Class {cls} ({cls_name})"] = count
                    movement_total += count
                row["Total"] = movement_total
                rows.append(row)

            # Approach subtotal row
            row = {"Approach": approach, "Movement": "SUBTOTAL"}
            approach_total = 0
            for cls in class_codes:
                cls_name = AUSTROADS_CLASSES.get(cls, {}).get("name", cls)
                cls_total = sum(
                    movements.get(m, {}).get(cls, 0) for m in MOVEMENT_ORDER
                )
                row[f"Class {cls} ({cls_name})"] = cls_total
                approach_total += cls_total
            row["Total"] = approach_total
            rows.append(row)

        if not rows:
            rows = [{"Approach": "No data", "Movement": "", "Total": 0}]

        pd.DataFrame(rows).to_excel(
            writer, index=False, sheet_name="TMC Summary"
        )

    def _write_tmc_interval(
        self,
        writer: pd.ExcelWriter,
        tmc_data: dict,
        od_matrix: ODMatrix,
    ) -> None:
        """TMC 15-min sheet: per-interval TMC breakdown."""
        class_codes = od_matrix.get_class_codes() or ["1"]
        rows = []

        for interval_key in sorted(tmc_data.keys()):
            interval_tmc = tmc_data[interval_key]
            for approach, movements in sorted(interval_tmc.items()):
                for movement in MOVEMENT_ORDER:
                    classes = movements.get(movement, {})
                    row = {
                        "Interval": interval_key,
                        "Approach": approach,
                        "Movement": movement,
                    }
                    movement_total = 0
                    for cls in class_codes:
                        cls_name = AUSTROADS_CLASSES.get(cls, {}).get(
                            "name", cls
                        )
                        count = classes.get(cls, 0)
                        row[f"Class {cls} ({cls_name})"] = count
                        movement_total += count
                    row["Total"] = movement_total
                    if movement_total > 0:
                        rows.append(row)

        if not rows:
            rows = [
                {
                    "Interval": "No data",
                    "Approach": "",
                    "Movement": "",
                    "Total": 0,
                }
            ]

        pd.DataFrame(rows).to_excel(
            writer, index=False, sheet_name="TMC 15-min"
        )

    def _write_od_matrix(
        self, writer: pd.ExcelWriter, od_matrix: ODMatrix
    ) -> None:
        """O-D Matrix sheet: full matrix summed across all classes."""
        df = od_matrix.to_dataframe()
        if df.empty:
            df = pd.DataFrame({"No data": [0]})
        df.to_excel(writer, sheet_name="O-D Matrix")

    def _write_od_by_class(
        self, writer: pd.ExcelWriter, od_matrix: ODMatrix
    ) -> None:
        """O-D by Class sheet: one section per vehicle class."""
        class_codes = od_matrix.get_class_codes()
        zone_names = od_matrix.get_zone_names()

        if not class_codes or not zone_names:
            pd.DataFrame({"No data": [0]}).to_excel(
                writer, index=False, sheet_name="O-D by Class"
            )
            return

        all_rows = []
        for cls in class_codes:
            cls_name = AUSTROADS_CLASSES.get(cls, {}).get("name", cls)
            # Header row for this class
            header_row = {"Zone": f"--- Class {cls}: {cls_name} ---"}
            for dest in zone_names:
                header_row[dest] = ""
            all_rows.append(header_row)

            # Matrix for this class
            cls_matrix = od_matrix.get_matrix_by_class(cls)
            for origin in zone_names:
                row = {"Zone": origin}
                for dest in zone_names:
                    row[dest] = cls_matrix.get(origin, {}).get(dest, 0)
                all_rows.append(row)

            # Blank separator
            all_rows.append({k: "" for k in ["Zone"] + zone_names})

        pd.DataFrame(all_rows).to_excel(
            writer, index=False, sheet_name="O-D by Class"
        )

    # ── Formatting ────────────────────────────────────────────────────

    def _format_workbook(
        self,
        filepath: str,
        job_config: JobConfig,
        od_matrix: ODMatrix,
    ) -> None:
        """Apply Matrix branding and report header to the workbook."""
        wb = load_workbook(filepath)

        # Format all sheets
        for ws in wb.worksheets:
            # Style header row
            for cell in ws[1]:
                cell.fill = _header_fill
                cell.font = _header_font
                cell.alignment = _center_align
                cell.border = _thin_border

            # Style data rows
            for row_idx, row in enumerate(
                ws.iter_rows(
                    min_row=2, max_row=ws.max_row, max_col=ws.max_column
                ),
                start=2,
            ):
                for cell in row:
                    cell.alignment = _center_align
                    cell.border = _thin_border
                    # Alternate row shading
                    if row_idx % 2 == 0:
                        cell.fill = _light_fill

            # Auto-width columns
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Add report header to TMC Summary sheet
        if "TMC Summary" in wb.sheetnames:
            ws = wb["TMC Summary"]
            header_lines = self._build_report_header(job_config, od_matrix)
            ws.insert_rows(1, len(header_lines) + 1)

            for i, (text, font) in enumerate(header_lines):
                cell = ws.cell(row=i + 1, column=1, value=text)
                cell.font = font
                cell.alignment = _left_align
                cell.border = Border()  # No border on header

        wb.save(filepath)

    def _build_report_header(
        self, job_config: JobConfig, od_matrix: ODMatrix
    ) -> list[tuple[str, Font]]:
        """Build report header lines with fonts."""
        title_font = Font(
            bold=True, size=16, color=_ACCENT_BLUE, name="Segoe UI"
        )
        info_font = Font(size=11, name="Segoe UI")
        accent_info = Font(size=11, color=_ACCENT_BLUE, name="Segoe UI")

        lines = [
            (
                "Matrix Intersection Turning Movement Count Report",
                title_font,
            ),
        ]

        if job_config.job_number or job_config.job_name:
            lines.append(
                (
                    f"Job: {job_config.job_number}  -  {job_config.job_name}",
                    info_font,
                )
            )

        if job_config.date_display:
            lines.append((f"Survey Date: {job_config.date_display}", info_font))

        if job_config.time_display:
            lines.append(
                (f"Survey Period: {job_config.time_display}", info_font)
            )

        total = od_matrix.get_total_count()
        lines.append((f"Total Vehicle Movements: {total}", accent_info))

        zone_names = od_matrix.get_zone_names()
        if zone_names:
            lines.append(
                (f"Zones: {', '.join(zone_names)}", info_font)
            )

        intervals = od_matrix.get_interval_keys()
        if intervals:
            lines.append(
                (
                    f"Time Range: {intervals[0]} to {intervals[-1]}",
                    info_font,
                )
            )

        return lines

    # ── Helpers ────────────────────────────────────────────────────────

    @staticmethod
    def _compute_total_tmc(
        tmc_data: dict,
    ) -> dict[str, dict[str, dict[str, int]]]:
        """Sum TMC data across all intervals.

        Returns:
            approach -> movement -> class_code -> count
        """
        totals: dict[str, dict[str, dict[str, int]]] = {}

        for interval_tmc in tmc_data.values():
            for approach, movements in interval_tmc.items():
                if approach not in totals:
                    totals[approach] = {}
                for movement, classes in movements.items():
                    if movement not in totals[approach]:
                        totals[approach][movement] = {}
                    for cls, count in classes.items():
                        totals[approach][movement][cls] = (
                            totals[approach][movement].get(cls, 0) + count
                        )

        return totals
