"""Excel export for ANPR results with survey details."""

import logging

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.common.survey_widget import SurveyInfo

logger = logging.getLogger(__name__)


def export_anpr_results(
    results: list[dict], survey: SurveyInfo, filepath: str
) -> None:
    """Export ANPR results to a formatted Excel file.

    Args:
        results: List of plate detection dicts. Supports both legacy format
                 (plate, time, confidence, valid, direction, video_file) and
                 vehicle-store format (vehicle_id, readings_count, user_corrected, etc.).
        survey: Survey details for the report header.
        filepath: Output .xlsx file path.
    """
    rows = []
    for r in results:
        row = {
            "Plate Number": r.get("plate", ""),
            "Video Time": r.get("time", r.get("real_time", "")),
            "Real Time": r.get("real_time", r.get("time", "")),
            "Confidence (%)": r.get("confidence", 0.0),
            "Valid Format": "Yes" if r.get("valid", r.get("is_valid")) else "No",
            "Direction": r.get("direction", ""),
            "Video File": r.get("video_file", ""),
        }
        # Include vehicle-store fields if present
        if "vehicle_id" in r:
            row["Vehicle ID"] = r["vehicle_id"]
        if "readings_count" in r:
            row["Readings"] = r["readings_count"]
        if "user_corrected" in r and r["user_corrected"]:
            row["User Corrected"] = r["user_corrected"]
        rows.append(row)

    # Determine columns based on whether vehicle data is present
    has_vehicle_data = any("vehicle_id" in r for r in results)
    if has_vehicle_data:
        columns = [
            "Vehicle ID", "Plate Number", "Video Time", "Real Time",
            "Confidence (%)", "Valid Format", "Direction", "Video File",
            "Readings", "User Corrected",
        ]
    else:
        columns = [
            "Plate Number", "Video Time", "Real Time",
            "Confidence (%)", "Valid Format", "Direction", "Video File",
        ]

    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=columns)
    else:
        # Reorder columns, dropping any that don't exist
        available = [c for c in columns if c in df.columns]
        df = df[available]

    df.to_excel(filepath, index=False, sheet_name="ANPR Results", engine="openpyxl")

    # Apply formatting
    wb = load_workbook(filepath)
    ws = wb.active

    header_fill = PatternFill(start_color="0f3460", end_color="0f3460", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin", color="2a2a4a"),
        right=Side(style="thin", color="2a2a4a"),
        top=Side(style="thin", color="2a2a4a"),
        bottom=Side(style="thin", color="2a2a4a"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # Auto-size columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # Add survey info header rows
    header_rows = [
        ("Matrix ANPR Data Extraction Report", Font(bold=True, size=14, color="e94560")),
    ]
    if survey.job_number or survey.job_name:
        header_rows.append((
            f"Job: {survey.job_number}  —  {survey.job_name}",
            Font(size=11),
        ))
    if survey.site_number or survey.site_name:
        header_rows.append((
            f"Site: {survey.site_number}  —  {survey.site_name}",
            Font(size=11),
        ))
    if survey.camera_number:
        header_rows.append((f"Camera: {survey.camera_number}", Font(size=11)))
    header_rows.append((f"Total Plates: {len(results)}", Font(size=11)))

    ws.insert_rows(1, len(header_rows) + 1)
    for i, (text, font) in enumerate(header_rows):
        ws[f"A{i + 1}"] = text
        ws[f"A{i + 1}"].font = font

    wb.save(filepath)
    logger.info("ANPR results exported to %s (%d plates)", filepath, len(results))
